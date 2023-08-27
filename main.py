from openpyxl import Workbook
from openpyxl import load_workbook
from scholarly import scholarly, ProxyGenerator 
import pandas as pd
from bs4 import BeautifulSoup
import requests, Proxy

FILENAME = "Publications_05.22.2023 (1).xlsx"
SHEETNAME = "Full papers"
cols = ["Professor", "Title", "Journal", "Year"]

def get_link(code): 
    return "http://scholar.google.com/scholar?hl=en&q=info:"+code+":scholar.google.com/&output=cite&scirp=0&hl=en"

def file_copy_path_generator(file_path=None): 
    cnt=1
    file_path_units = file_path.split('')
    while(True):
        new_file_name = "("+str(cnt)+")"+file_path
        try: 
            with open(new_file_name, "r"): 
                pass
        except: 
            return new_file_name 

#sheet is defaulted to 1 
def process_excel(file): 
    max_row = 1; 
    wb = load_workbook(filename=file)
    ws = wb.active
    while(ws[str('B')+str(max_row)].value!=None): 
        max_row=max_row+1
    return [wb, ws, max_row] 

#task 1
def clean_duplicates(file=None):
    if file is None: 
        file = input("Enter File Path: ")
    ls = process_excel(file) 
    wb=ls[0], ws=ls[1], max_row=ls[2]
    #finding duplicates using counter array
    dup = []
    for i in range(0, max_row): 
        dup.append(0)

    for i in range(1, max_row):
        if(dup[i]!=0):
            continue
        title = ws[str('B')+str(i)].value
        journal = ws[str('C')+str(i)].value
        dup[i]=i
        for j in range(i+1, max_row):
            tl = ws[str('B')+str(j)].value
            jl = ws[str('C')+str(j)].value
            if(tl==title and jl==journal): 
                dup[j]=i 

    action_dup_row = []
    #presenting duplciates
    for i in range(1, max_row):  
        dup_row = [i]
        for j in range(i+1, max_row): 
            if(dup[j]==dup[i]): 
                dup_row.append(j)
                action_dup_row.append(j) 
        if(len(dup_row)>1): 
            print("Duplicates Found: ")
            for k in range(len(dup_row)):
                print("ROW "+str(dup_row[k]))
            print("\n")

    action_dup_row.sort(reverse=True)
    option = input("Clean Duplicates(y/n): ")
    if(option=='y'): 
        for i in range(len(action_dup_row)):
            ws.delete_rows(action_dup_row[i], 1)

    #
    wb.save(file+"")

#task 2
def sort_list(file_path=None, sheet="Full papers", opt1=None, opt2=None): 
    if file_path is None: 
        file_path = input("Enter File Path: ") 
`   df = pd.read_excel(file_path)
    for i in range(len(cols)):
        print("("+str(i+1)+")", cols[i], " ")  
    if(opt1 is None):
        opt1 = int(input("Sort By: "))
    while(opt1<1 or opt1>4): 
        print("ERROR")
        opt1 = input("Sort By: ")
    if(opt2 is None): 
        opt2 = input("ascending(a) / descending(d): ")
    while(opt2!='a' and opt2!='d'): 
        print("ERROR")
        opt2 = input("ascending(a) / descending(d): ")
    opt1 -= 1
    opt2 = True if(opt2=='a') else False
    print(opt2, cols[opt1])

    df.sort_values(by=[cols[opt1]], ascending=opt2, inplace=True)
    df.to_excel(file_path, sheet_name=sheet, index=False)
    print("SUCCESSED")`


    
#base method for task 3/4/5
#return a dateframe containign all of a author's publication in a time range
#all pubs are unique in the df
def get_author_pubs(author_name=None, year_l_bound=0, year_r_bound=9999):
    if author_name is None: 
        author_name = input("Enter Author Name: ")    
    pg = ProxyGenerator()
    success = pg.FreeProxies()
    scholarly.use_proxy(pg)
    df = pd.DataFrame(columns=['Professor','Title','Journal', 'Year', 'Citations', 'php1', 'php2', 'cite'])      
    search_query = scholarly.search_author(author_name)
    first_author_result = next(search_query)
    author = scholarly.fill(first_author_result)
    for pub in author['publications']:
        scholarly.fill(pub)
        pub_title = pub['bib']['title']
        string = pub['url_related_articles'].split(':')
        code = ""
        for i in range(len(string)): 
            if(string[i]=='scholar.google.com/'): 
                code = string[i-1] 
        link = get_link(code)                    
        pub_journal = pub['bib']['citation']
        pub_date = pub['bib']['pub_year']
        if(pub_date>=year_l_bound and pub_date<=year_r_bound):
            pub_citation = pub['num_citations']
            php_text = "<li>"+pub_title+"</li>"
            php_text_2 = "<li>"+pub_title+" in "+ pub_journal+ "</li>"
            html = requests.get(link).text
            soup = BeautifulSoup(html, "lxml")
            apa_cit = "<li>"+soup.find('th', string="APA").find_parent().find('div').text+"</li>"
            new_data = {'Professor': author_name, 
                        'Title': pub_title, 'Journal': pub_journal, 
                        'Year': pub_date, 'Citations': pub_citation, 
                        'php1': php_text, 'php2':php_text_2, 'cite': apa_cit}
            df.loc[len(df.index)] = new_data
    return df

#return a list with all the indexes of a author's publication in the excel file
def search_author_pubs(file_path=None, sheet="Full papers", author_name=None, year_l_bound=0, year_r_bound=9999): 
    if file_path is None: 
        file_path = input("Enter File Path")
    if author_name is None: 
        author_name = input("Enter Author Name: ")
    df = pd.read_excel(file_path, sheet)
    author_names=df['Author'], years=df['year']
    ret_list = []
    for i in range(len(df.index)):
        if(author_name==author_names[i] and years[i]>=year_l_bound and years[i]<=year_r_bound):
            ret_list.append(i)
    return ret_list
    
#task 3
# update all authors 
def update_list(file_path=None, sheet="Full papers"): 
    if(file_path is None): 
        file_path = input("Enter File Path: ")
    df = pd.read_excel(file_path)
    serie = list(df['Author'])
    serie = set(serie)
    for i in range(len(serie)): 
        new_df = get_author_pubs(serie[i])
        df = pd.concat([df, new_df], ignore_index=True)
    df.to_excel(file_path, sheet_name=sheet)

#task 4
def modify_list(file_path=None, sheet="Full papers", year_l_bound=0, year_r_bound=9999): 
    if file_path is None: 
        file_path = input("Enter File Path: ")
    df = pd.read_excel(file_path)
    author_names = {}
    serie = df['Author']
    max_row = df.axes[0].stop
    for i in range(max_row): 
        author_names[serie[i]]=True
    #check to see if author name exists
    author_name_input = input("Enter Author Name: ")
    if author_names[author_name_input] is True: 
        year_l_bound = int(input("Year(left bound): "))
        year_r_bound = int(input("Year(right bound): "))
        search_result = []
        for i in range(max_row): 
            if serie[i]==author_names:
                year = df.iloc(i)['Year']
                if(year>=year_l_bound and year<=year_r_bound): 
                    search_result.append(i)
        if(len(search_result)==0): 
            print("No Result")
            return
        print("INDEX    ROW")
        for i in range(len(search_result)):
            print(i+1,"       ",search_result[i])
        opt = int(input("Enter INDEX to delete (0 to quit): "))
        while(opt!=0 and len(search_result)!=0):
            opt = int(input("Enter INDEX to delete (0 to quit): "))
            try: 
                df.drop(labels=[search_result[opt-1]], axis=0, inplace=False)
            except: 
                print("Error")
        df.to_excel(file_path, sheet_name=sheet, index=False)
    else: 
        print("Author not Found")

#task 5
def add_author(file_path=None, sheet="Full papers", author_name=None): 
    if(file_path is None): 
        file_path = input("Enter File Path")
    if(author_name is None): 
        author_name = input("Enter Author Name") 
    df = pd.read_excel(file_path)
    new_df = get_author_pubs(author_name)
    print(new_df)
    opt = input("Add to excel?(yes/no): ")
    if(opt=='yes'): 
        df = pd.concat([df, new_df], ignore_index=True)
    df.to_excel(file_path, sheet_name=sheet)

#task 6 
def generate_php(file_path=None): 
    if file_path is None: 
        file_path = input("file_path: ")
    sort_list(file_path, 0, 4, 'd')
    df = pd.read_excel(file_path)
    years_serie = df['Year']
    cite_serie = df['Cite']
    cur_year = years_serie[0]
    with open("published-work.php", 'w') as f:
        f.write("")
    with open("published-work.php", 'a') as f: 
        f.write("<div id=\"contentContainer\"> \n<div id=\"content\"> \n<h1>Published Work</h1> \n<!--p>Individual lists of publications by CPCC faculty.</p-->")
        f.write("\n<h2>"+str(years_serie[0])+"</h2>")
        f.write("\n<ul>")
        for i in range(len(df.index)): 
            if(years_serie[i] != cur_year): 
                f.write("\n</ul>")
                f.write("\n<h2>"+str(years_serie[i])+"</h2>")
                f.write("<ul>")
                cur_year = years_serie[i]
            f.write("\n"+cite_serie[i])
        f.write("\n</ul>")
            
def ui(): 
    funct=[clean_duplicates,sort_list,modify_list,add_author,generate_php]
    file_path=input("Enter File Path: ")
    try: 
        with open(file_path, 'r') as file: 
            print("Read Successfully")
    except FileNotFoundError: 
        print("File Not Found")
        return    
    
    #) path may not be in current directory
    file_copy_path = file_copy_path_generator(file_path)
    while(True): 
        print("1)clean duplicates  2)sort entries  3)update author")
        print("4)modify entries  5)add author  6)generate php file")
        opt = int(input("Enter Operation: "))
        while(opt<1 or opt>6): 
            print("Error")
            opt = int(input("Enter Operation: "))
        funct[opt-1](file_copy_path)

if __name__=="__main__":
